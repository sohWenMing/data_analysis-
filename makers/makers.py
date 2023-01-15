import openpyxl
from openpyxl import Workbook
import pandas as pd
import numpy as np
import datetime as dt

dataset = input('Enter CSV filename \n (filename must have no spaces, please include .csv extension): ')
output = input('Enter Excel filename to save to \n (filename must have no spaces, please include .xlsx extension) : ')
"""
allows user to input CSV filename into script, for ease of use CSV should be in 
same folder as script
"""
data = pd.read_csv(dataset)
# reads input csv into variable data, change later
month_dict = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec"
}
#Month dictionary used later to map text representations of months

# Setting up of datatype params

# ##Year and Month##
data['Date'] = pd.to_datetime(data['Date'], dayfirst=True)
# changes all data in "Date" into datetime dtype
data['Month'] = data['Date'].dt.month
# extracts month value as integer from datetime stored in "Date" column
data['Month'] = data['Month'].map(month_dict)
# remaps month names to take over integer values
data['Year'] = data['Date'].dt.year
data['Year_Month'] = data['Year'].astype(str) + " " + data['Month']
# builds Year_Month Column so that each year in each month can be handled individually

# ## Generating Day of Week##
data['Day'] = data['Date'].dt.dayofweek

# ##Extracting of Hours of Day from 'Time' Column - original is a string with format
#  'dd:dd - dd:dd' - extraction of first characters to make up for edge cases where
# booking was more than 1 hour. As these are outliers, we will still count each booking
# on a one hour basis
data['Time'] = data['Time'].str[:5]
data['Time'] = pd.to_datetime(data['Time'])
data['Time'] = data['Time'].dt.hour

# #### Extracting Days of Week from Datetime ####

day_dict = {
    0: 'Monday',
    1: 'Tuesday',
    2: 'Wednesday',
    3: 'Thursday',
    4: 'Friday',
    5: 'Saturday',
    6: 'Sunday'
}
# Dictionary used to remap text based representation of days of week as .dayofweek
# extracts as an integer value
data['Day'] = data['Day'].map(day_dict)
data = data[data['Day'] != "Monday"]
# remove all Mondays

time_list = sorted(data['Time'].unique())
# list of all hours sorted, to be used to iterate to get hourly timings later
hour_dict = {
    9: '9am',
    10: '10am',
    11: '11am',
    12: '12 noon',
    13: '1pm',
    14: '2pm',
    15: '3pm',
    16: '4pm',
    17: '5pm',
    18: '6pm'
}

# used to convert hour in time_list to a readable time format

# ###removal of erroneous entries - entries made by principals of business###
data = data[(data['Client name'] != 'Amz') & \
            (data['Client name'] != "Aisyah Zainalabidin")]

if len(data[data['Client name'] == "Aisyah Zainalabidin"]) == 0 and \
        len(data[data['Client name'] == 'Amz']) == 0:
    print('Amz and Aisyah entries removed')
else:
    print('Error in removing principal\'s entries')
# checking for errors

# ### removal of cancellation entries ###
data = data[data['Is cancelled'] != "Yes"]
# removes all rows where 'Is cancelled' has a "Yes" entry

drop_more_cancel = data[(data['Cancellation type'] == 'nopayment_cancel') |
                        (data['Cancellation type'] == 'Cancelled')].index
data.drop(drop_more_cancel, inplace=True)
# removes all rows where cancellation type is "nopayment_cancel" or "Cancelled"

if len(data[data['Is cancelled'] == 'Yes']) == 0 and \
    len(data[data['Cancellation type'] == 'nopayment_cancel']) == 0 and \
        len(data[data['Cancellation type'] == 'Cancelled']) == 0:
    print('All cancelled entries removed')
else:
    print('Error in removing cancelled entries')
# checking for errors

# Data Manipulation and input into openpyxl
wb = openpyxl.Workbook()
dest_filename = output

# iterates through all possible year/month combinations in dataset
for year_month in data['Year_Month'].unique():
    print('Processing:', year_month)
    year_month_sheet = year_month
    wb.create_sheet(year_month_sheet)
    sheet = wb[year_month_sheet]
    # generates a sheet in excel file, name based on the year/month being iterated over

    # ####calculating of best customers####
    best_customer = data[data['Year_Month'] == year_month]['Client name'].value_counts().head(10)
    # creates a series of the 10 best customers and their hours logged for the month
    best_customer.rename("Best Customer Hours")
    customer_name = best_customer.index.values
    # extracts index names from best_customer series into iterable list customer_name
    best_customer_dict = {}
    for name, hour in zip(customer_name, best_customer):
        best_customer_dict[name] = hour
    # print(best_customer_dict)
    # use openpyxl to write best customer details to sheet
    for i in range(0, len(customer_name)):
        customer = customer_name[i]
        sheet.cell(row=i+3, column=1).value = customer
        sheet.cell(row=i+3, column=2).value = best_customer_dict[customer]
        sheet.cell(row=1, column=1).value = "Best Customers Of the Month"
        sheet.cell(row=2, column=1).value = 'Customer'
        sheet.cell(row=2, column=2).value = 'Hours'
    sheet.merge_cells('A1:B1')

        # calculating average hours per day
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday',
            'Friday', 'Saturday', 'Sunday']
    daily_hours = {}
    for day in days:
        month_day = data[(data["Year_Month"] == year_month) & (data['Day'] == day)]
        # extracts all entries where year_month is the correct year month for the day being iterated over
        if len(month_day['Date'].unique()) != 0:  # if days are logged where there are
            # the specific day in the Year_Month
            daily_hours[day] = round(len(month_day) / len(
                    month_day['Date'].unique()), 2)
            # divide the number of entries by the unique number of dates, to get hours per day that was logged
        day_keys = list(daily_hours.keys())
        # extract keys of daily_hours dictionary to an iterable list
    for i in range(0, len(day_keys)):
        day = day_keys[i]
        sheet.cell(row=i+3, column=6).value = day
        sheet.cell(row=i+3, column=7).value = daily_hours[day]
        sheet.cell(row=1, column=6).value = "Hours By Day"
        sheet.cell(row=2, column=6).value = "Day"
        sheet.cell(row=2, column=7).value = "Average Hours"
    sheet.merge_cells('F1:G1')
    # print(daily_hours)

    customer_per_hour_dict = {}
    for hour in time_list:
        correct_hours = data[(data['Year_Month'] == year_month) &
                            (data['Time'] == hour)]
        if len(correct_hours['Date'].unique()) != 0:
            clock_hour = hour_dict[hour]
            customer_per_hour_dict[clock_hour] = round(len(correct_hours) / len(
                correct_hours['Date'].unique()), 2)
    per_hour_keys = list(customer_per_hour_dict.keys())
    # extract all keys in customer_per_hour_dict dictionary
    for i in range(0, len(per_hour_keys)):
        time = per_hour_keys[i]
        sheet.cell(row=i+3, column=10).value = time
        sheet.cell(row=i+3, column=11).value = customer_per_hour_dict[time]
        sheet.cell(row=1, column=10).value = "Customers Per Hour"
        sheet.cell(row=2, column=10).value = 'Time'
        sheet.cell(row=2, column=11).value = "Number of Customers"
    sheet.merge_cells('J1:K1')
    # print(customer_per_hour_dict)

    #
    column_list = ['A', 'B', 'F', 'G', 'J', 'K']
    for column in column_list:
        sheet.column_dimensions[column].width = 20

del wb['Sheet']
wb.save(dest_filename)
wb.close()
print(data.info())


# calculating hourly volume









